"""CattleEyeView loading, validation, and training-data preparation."""

from __future__ import annotations

import csv
import json
import math
import os
import shutil
from collections.abc import Iterator
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

EXPECTED_FRAMES = 30_703
EXPECTED_SEQUENCES = 14
EXPECTED_TRACKS = 753
KEYPOINT_COUNT = 24
YOLO_TASKS = ("detect", "detect_head", "pose", "segment")
SPLITS = ("train", "val", "test")
KEYPOINT_NAMES = (
    "head",
    "nose",
    "eyeL",
    "eyeR",
    "earbaseL",
    "eartipL",
    "earbaseR",
    "eartipR",
    "neck",
    "withers",
    "elbowFL",
    "kneeFL",
    "pawFL",
    "elbowFR",
    "kneeFR",
    "pawFR",
    "elbowBL",
    "kneeBL",
    "pawBL",
    "elbowBR",
    "kneeBR",
    "pawBR",
    "tailbase",
    "tailend",
)
FLIP_INDEX = (
    0,
    1,
    3,
    2,
    6,
    7,
    4,
    5,
    8,
    9,
    13,
    14,
    15,
    10,
    11,
    12,
    19,
    20,
    21,
    16,
    17,
    18,
    22,
    23,
)


class DatasetValidationError(ValueError):
    """Raised when CattleEyeView cannot be safely loaded or converted."""


@dataclass(frozen=True)
class CattleEyeViewPaths:
    """Canonical paths within an extracted CattleEyeView release."""

    root: Path

    @classmethod
    def from_root(cls, root: Path | str) -> CattleEyeViewPaths:
        return cls(Path(root).expanduser().resolve())

    @property
    def images(self) -> Path:
        return self.root / "images"

    @property
    def videos(self) -> Path:
        return self.root / "videos"

    @property
    def annotations(self) -> Path:
        return self.root / "annotation"

    @property
    def metadata_workbook(self) -> Path:
        return self.annotations / "metadata and count.xlsx"

    def tracking_json(self, split: str) -> Path:
        if split not in {"train", "test"}:
            raise ValueError(f"Tracking split must be train or test, got {split!r}")
        return self.annotations / "pose_COCO" / f"coco_track_{split}.json"

    def task_labels(self, task: str) -> Path:
        if task not in YOLO_TASKS:
            raise ValueError(f"Unsupported YOLO task: {task}")
        return self.annotations / task / "labels"

    def image_for(self, file_name: str) -> Path:
        return self.images.joinpath(*PurePosixPath(file_name).parts)


@dataclass(frozen=True)
class SequenceMetadata:
    """Ground-truth metadata and crossing count for one video sequence."""

    video_filename: str
    capture_date: str
    start_time: str
    end_time: str
    frames: int
    split: str
    count: int


@dataclass(frozen=True)
class TrackingRecord:
    """One normalized, tracked cattle annotation in one frame."""

    split: str
    sequence: str
    frame: str
    frame_index: int
    image: str
    video: str
    track_id: int
    track_uid: str
    bbox_xywh: list[float]
    head_bbox_xywh: list[float] | None
    keypoints: list[float]
    visible_keypoints: int
    occluded: bool
    truncated: bool
    annotation_id: int
    source_image_id: int


@dataclass
class ValidationReport:
    """Machine-readable validation result."""

    root: str
    summary: dict[str, Any] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.errors

    def to_dict(self) -> dict[str, Any]:
        return {
            "root": self.root,
            "valid": self.valid,
            "summary": self.summary,
            "errors": self.errors,
            "warnings": self.warnings,
        }


@dataclass(frozen=True)
class PreparedDataset:
    """Summary of a generated Ultralytics-compatible dataset."""

    task: str
    output: str
    link_mode: str
    split_images: dict[str, int]
    yaml: str
    manifest: str


def _iso_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (date, time)):
        return value.isoformat()
    return str(value)


def load_sequence_metadata(root: Path | str) -> list[SequenceMetadata]:
    """Load official video splits and crossing counts from the workbook."""
    paths = CattleEyeViewPaths.from_root(root)
    if not paths.metadata_workbook.is_file():
        raise DatasetValidationError(
            f"Count workbook not found: {paths.metadata_workbook}"
        )

    workbook = load_workbook(
        paths.metadata_workbook,
        read_only=True,
        data_only=True,
    )
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = tuple(next(rows))
        expected = (
            "video_filename",
            "Date",
            "Time Start",
            "Time End",
            "Frames",
            "Split",
            "Count",
        )
        if headers != expected:
            raise DatasetValidationError(
                f"Unexpected count workbook columns: {headers!r}"
            )

        metadata = []
        for row in rows:
            if not row or not row[0]:
                continue
            metadata.append(
                SequenceMetadata(
                    video_filename=str(row[0]),
                    capture_date=_iso_value(row[1]),
                    start_time=_iso_value(row[2]),
                    end_time=_iso_value(row[3]),
                    frames=int(row[4]),
                    split=str(row[5]).lower(),
                    count=int(row[6]),
                )
            )
        return metadata
    finally:
        workbook.close()


def _load_tracking_document(paths: CattleEyeViewPaths, split: str) -> dict[str, Any]:
    source = paths.tracking_json(split)
    if not source.is_file():
        raise DatasetValidationError(f"Tracking annotations not found: {source}")
    with source.open("r", encoding="utf-8") as file:
        return json.load(file)


def _optional_box(value: Any) -> list[float] | None:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if not isinstance(value, list) or len(value) != 4:
        return None
    box = [float(item) for item in value]
    return box if all(math.isfinite(item) for item in box) else None


def iter_tracking_records(
    root: Path | str,
    splits: tuple[str, ...] = ("train", "test"),
) -> Iterator[TrackingRecord]:
    """Yield normalized records, using file_name instead of unreliable image_id."""
    paths = CattleEyeViewPaths.from_root(root)
    for split in splits:
        document = _load_tracking_document(paths, split)
        images_by_name = {
            str(image["file_name"]): image for image in document.get("images", [])
        }
        for annotation in document.get("annotations", []):
            file_name = str(annotation["file_name"])
            image = images_by_name.get(file_name)
            if image is None:
                raise DatasetValidationError(
                    f"Annotation references unknown image: {file_name}"
                )
            posix_path = PurePosixPath(file_name)
            track_id = int(annotation["instance_id"])
            keypoints = [float(value) for value in annotation.get("keypoints", [])]
            yield TrackingRecord(
                split=split,
                sequence=posix_path.parent.name,
                frame=posix_path.name,
                frame_index=int(image.get("frame_id", int(posix_path.stem) - 1)),
                image=(PurePosixPath("images") / posix_path).as_posix(),
                video=(PurePosixPath("videos") / posix_path.parent.name).as_posix(),
                track_id=track_id,
                track_uid=f"{posix_path.parent.name}:{track_id}",
                bbox_xywh=[float(value) for value in annotation["bbox"]],
                head_bbox_xywh=_optional_box(annotation.get("bbox_head")),
                keypoints=keypoints,
                visible_keypoints=int(annotation.get("num_keypoints", 0)),
                occluded=bool(annotation.get("occluded", False)),
                truncated=bool(annotation.get("truncated", False)),
                annotation_id=int(annotation["id"]),
                source_image_id=int(annotation["image_id"]),
            )


def _label_schema_error(task: str, tokens: list[str]) -> str | None:
    if not tokens:
        return None
    try:
        values = [float(token) for token in tokens]
    except ValueError:
        return "contains a non-numeric token"
    if values[0] != 0.0:
        return "uses a class ID other than 0"

    if task in {"detect", "detect_head"} and len(values) != 5:
        return f"requires 5 values, found {len(values)}"
    if task == "pose" and len(values) != 5 + KEYPOINT_COUNT * 3:
        return f"requires 77 values, found {len(values)}"
    if task == "segment" and (len(values) < 7 or len(values) % 2 == 0):
        return "requires a class ID followed by at least 3 x/y points"

    normalized = values[1:5] if task != "segment" else values[1:]
    if task == "pose":
        normalized = values[1:5]
        for index in range(5, len(values), 3):
            normalized.extend(values[index : index + 2])
            if values[index + 2] not in {0.0, 1.0, 2.0}:
                return "contains an invalid keypoint visibility value"
    if any(value < 0.0 or value > 1.0 for value in normalized):
        return "contains a normalized coordinate outside [0, 1]"
    return None


def validate_dataset(
    root: Path | str,
    *,
    validate_labels: bool = True,
    label_sample_limit: int | None = 25,
) -> ValidationReport:
    """Validate files, splits, tracking identities, counts, and YOLO labels."""
    paths = CattleEyeViewPaths.from_root(root)
    report = ValidationReport(root=str(paths.root))
    for required in (paths.images, paths.videos, paths.annotations):
        if not required.is_dir():
            report.errors.append(f"Required directory is missing: {required}")
    if report.errors:
        return report

    try:
        metadata = load_sequence_metadata(paths.root)
    except (DatasetValidationError, OSError, ValueError) as error:
        report.errors.append(str(error))
        return report

    frame_counts: dict[str, int] = {}
    available_images: set[str] = set()
    missing_videos: list[str] = []
    for sequence in metadata:
        sequence_dir = paths.images / sequence.video_filename
        frame_paths = tuple(sequence_dir.glob("*.jpg")) if sequence_dir.is_dir() else ()
        frame_count = len(frame_paths)
        available_images.update(
            f"{sequence.video_filename}/{frame_path.name}"
            for frame_path in frame_paths
        )
        frame_counts[sequence.video_filename] = frame_count
        if frame_count != sequence.frames:
            report.errors.append(
                f"{sequence.video_filename}: workbook expects {sequence.frames} "
                f"frames, found {frame_count}"
            )
        if not (paths.videos / sequence.video_filename).is_file():
            missing_videos.append(sequence.video_filename)
    if missing_videos:
        report.errors.append(f"Missing videos: {', '.join(missing_videos)}")

    tracking_annotations = 0
    tracks: set[tuple[str, int]] = set()
    missing_tracking_images = 0
    malformed_boxes = 0
    image_id_mismatches = 0
    split_sequences: dict[str, set[str]] = {"train": set(), "test": set()}
    for split in ("train", "test"):
        try:
            document = _load_tracking_document(paths, split)
        except (DatasetValidationError, OSError, json.JSONDecodeError) as error:
            report.errors.append(str(error))
            continue
        images_by_name = {
            str(image["file_name"]): image for image in document.get("images", [])
        }
        images_by_id = {
            int(image["id"]): image for image in document.get("images", [])
        }
        for annotation in document.get("annotations", []):
            tracking_annotations += 1
            file_name = str(annotation.get("file_name", ""))
            sequence = PurePosixPath(file_name).parent.name
            split_sequences[split].add(sequence)
            tracks.add((sequence, int(annotation["instance_id"])))
            if (
                file_name not in images_by_name
                or file_name not in available_images
            ):
                missing_tracking_images += 1
            source_image = images_by_id.get(int(annotation.get("image_id", -1)))
            if source_image and source_image.get("file_name") != file_name:
                image_id_mismatches += 1
            bbox = annotation.get("bbox")
            if (
                not isinstance(bbox, list)
                or len(bbox) != 4
                or float(bbox[2]) <= 0
                or float(bbox[3]) <= 0
            ):
                malformed_boxes += 1
    if missing_tracking_images:
        report.errors.append(
            f"{missing_tracking_images} tracking annotations reference missing images"
        )
    if malformed_boxes:
        report.errors.append(f"{malformed_boxes} tracking boxes are malformed")
    if image_id_mismatches:
        report.warnings.append(
            f"The source has {image_id_mismatches} inconsistent COCO image_id "
            "references; normalized records use file_name instead."
        )

    task_labels: dict[str, dict[str, int]] = {}
    checked_label_files = 0
    if validate_labels:
        invalid_labels: list[str] = []
        missing_label_images: list[str] = []
        for task in YOLO_TASKS:
            task_labels[task] = {}
            label_root = paths.task_labels(task)
            for split in SPLITS:
                split_root = label_root / split
                label_files = sorted(split_root.glob("*/*.txt"))
                task_labels[task][split] = len(label_files)
                labels_to_check = (
                    label_files
                    if label_sample_limit is None
                    else label_files[:label_sample_limit]
                )
                for label_path in label_files:
                    sequence = label_path.parent.name
                    file_name = f"{sequence}/{label_path.stem}.jpg"
                    if (
                        file_name not in available_images
                        and len(missing_label_images) < 10
                    ):
                        missing_label_images.append(str(label_path))
                for label_path in labels_to_check:
                    checked_label_files += 1
                    with label_path.open("r", encoding="utf-8") as label_file:
                        for line_number, line in enumerate(label_file, start=1):
                            error = _label_schema_error(task, line.split())
                            if error and len(invalid_labels) < 10:
                                invalid_labels.append(
                                    f"{label_path}:{line_number} {error}"
                                )
            if not label_root.is_dir():
                report.errors.append(f"Missing {task} labels: {label_root}")
        if missing_label_images:
            report.errors.append(
                "Labels with missing source images (first 10): "
                + "; ".join(missing_label_images)
            )
        if invalid_labels:
            report.errors.append(
                "Invalid native labels (first 10): " + "; ".join(invalid_labels)
            )

    total_frames = sum(frame_counts.values())
    total_count = sum(sequence.count for sequence in metadata)
    report.summary = {
        "sequences": len(metadata),
        "frames": total_frames,
        "videos": sum(
            (paths.videos / sequence.video_filename).is_file()
            for sequence in metadata
        ),
        "tracking_annotations": tracking_annotations,
        "unique_tracks": len(tracks),
        "ground_truth_crossings": total_count,
        "sequence_splits": {
            split: sorted(sequences) for split, sequences in split_sequences.items()
        },
        "task_label_files": task_labels,
        "label_files_content_checked": checked_label_files,
    }
    if len(metadata) != EXPECTED_SEQUENCES:
        report.warnings.append(
            f"Expected {EXPECTED_SEQUENCES} sequences, found {len(metadata)}"
        )
    if total_frames != EXPECTED_FRAMES:
        report.warnings.append(
            f"Expected {EXPECTED_FRAMES} frames, found {total_frames}"
        )
    if len(tracks) != EXPECTED_TRACKS:
        report.warnings.append(
            f"Expected {EXPECTED_TRACKS} tracked cattle, found {len(tracks)}"
        )
    return report


def export_tracking_manifest(
    root: Path | str,
    output_dir: Path | str,
) -> dict[str, str]:
    """Export normalized JSONL tracking records and sequence count metadata."""
    paths = CattleEyeViewPaths.from_root(root)
    output = Path(output_dir).expanduser().resolve()
    output.mkdir(parents=True, exist_ok=True)
    manifest_path = output / "tracking_manifest.jsonl"
    metadata_path = output / "sequences.json"
    report_path = output / "validation_report.json"

    report = validate_dataset(paths.root)
    if not report.valid:
        raise DatasetValidationError(
            "Dataset validation failed: " + "; ".join(report.errors)
        )

    with manifest_path.open("w", encoding="utf-8", newline="\n") as file:
        for record in iter_tracking_records(paths.root):
            file.write(json.dumps(asdict(record), allow_nan=False) + "\n")
    with metadata_path.open("w", encoding="utf-8", newline="\n") as file:
        json.dump(
            [asdict(item) for item in load_sequence_metadata(paths.root)],
            file,
            indent=2,
        )
        file.write("\n")
    with report_path.open("w", encoding="utf-8", newline="\n") as file:
        json.dump(report.to_dict(), file, indent=2)
        file.write("\n")
    return {
        "manifest": str(manifest_path),
        "sequences": str(metadata_path),
        "validation_report": str(report_path),
    }


def _create_link(source: Path, destination: Path, mode: str) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        if destination.stat().st_size != source.stat().st_size:
            raise DatasetValidationError(
                f"Existing destination differs from source: {destination}"
            )
        return
    if mode == "hardlink":
        os.link(source, destination)
    elif mode == "copy":
        shutil.copy2(source, destination)
    else:
        raise ValueError(f"Unsupported link mode: {mode}")


def _write_dataset_yaml(task: str, output: Path) -> Path:
    yaml_path = output / "dataset.yaml"
    escaped_root = output.as_posix().replace('"', '\\"')
    lines = [
        f'path: "{escaped_root}"',
        "train: images/train",
        "val: images/val",
        "test: images/test",
        "names:",
        "  0: cow",
    ]
    if task == "pose":
        lines.extend(
            [
                f"kpt_shape: [{KEYPOINT_COUNT}, 3]",
                "flip_idx: [" + ", ".join(map(str, FLIP_INDEX)) + "]",
            ]
        )
    yaml_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return yaml_path


def prepare_yolo_dataset(
    root: Path | str,
    task: str,
    output_dir: Path | str | None = None,
    *,
    link_mode: str = "hardlink",
) -> PreparedDataset:
    """Create a correct Ultralytics layout from canonical images and labels."""
    if task not in YOLO_TASKS:
        raise ValueError(f"Task must be one of {', '.join(YOLO_TASKS)}")
    if link_mode not in {"hardlink", "copy"}:
        raise ValueError("link_mode must be hardlink or copy")
    paths = CattleEyeViewPaths.from_root(root)
    output = (
        Path(output_dir).expanduser().resolve()
        if output_dir is not None
        else paths.root / "prepared" / task
    )

    pairs: list[tuple[str, str, str, Path, Path]] = []
    missing_images: list[Path] = []
    for split in SPLITS:
        for label in sorted((paths.task_labels(task) / split).glob("*/*.txt")):
            sequence = label.parent.name
            frame = f"{label.stem}.jpg"
            image = paths.images / sequence / frame
            if not image.is_file():
                missing_images.append(image)
            pairs.append((split, sequence, frame, image, label))
    if missing_images:
        examples = ", ".join(map(str, missing_images[:10]))
        raise DatasetValidationError(
            f"Cannot prepare {task}: {len(missing_images)} images missing; {examples}"
        )

    manifest_path = output / "manifest.csv"
    output.mkdir(parents=True, exist_ok=True)
    split_counts = {split: 0 for split in SPLITS}
    with manifest_path.open("w", encoding="utf-8", newline="") as manifest_file:
        writer = csv.DictWriter(
            manifest_file,
            fieldnames=("split", "sequence", "frame", "image", "label"),
        )
        writer.writeheader()
        for split, sequence, frame, image, label in pairs:
            output_image = output / "images" / split / sequence / frame
            output_label = output / "labels" / split / sequence / label.name
            _create_link(image, output_image, link_mode)
            _create_link(label, output_label, link_mode)
            split_counts[split] += 1
            writer.writerow(
                {
                    "split": split,
                    "sequence": sequence,
                    "frame": frame,
                    "image": output_image.relative_to(output).as_posix(),
                    "label": output_label.relative_to(output).as_posix(),
                }
            )

    yaml_path = _write_dataset_yaml(task, output)
    return PreparedDataset(
        task=task,
        output=str(output),
        link_mode=link_mode,
        split_images=split_counts,
        yaml=str(yaml_path),
        manifest=str(manifest_path),
    )
